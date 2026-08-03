#!/usr/bin/env python3
"""Rafraîchit le Factbook depuis les JSON pays de la CIA (multi-plateforme).

Chaîne complète, équivalente à update_factbook.ps1 mais exécutable sous
Linux/macOS/Windows :

    JSON pays (factbook/factbook.json)
      -> cia-facbook-pwa/FACTBOOK_FMCG_2026_TOTALEMENT_A_JOUR.xlsx
      -> cia-facbook-pwa/factbook_data.json
      -> src/data/factbook.json   (via scripts/build_factbook.py)

Ajoute les colonnes Main_Cluster (NG / RoSSA) et Sub_Cluster (NG / WA / EESA).

    pip install pandas openpyxl
    python3 scripts/update_factbook.py
"""

from __future__ import annotations

import json
import re
import subprocess
from pathlib import Path

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

from build_factbook import WA_COUNTRIES, main as build_json

ROOT = Path(__file__).resolve().parents[1]
PWA = ROOT / "cia-facbook-pwa"
SOURCE_XLSX = PWA / "FMCG_Factbook_Optimized_For_PWA.xlsx"
TARGET_XLSX = PWA / "FACTBOOK_FMCG_2026_TOTALEMENT_A_JOUR.xlsx"
TARGET_JSON = PWA / "factbook_data.json"
CLONE = PWA / "factbook.json"
REPO = "https://github.com/factbook/factbook.json.git"

# Code CIA de chaque marché (nom de fichier dans factbook.json/africa/).
COUNTRY_CODES = {
    "Nigeria": "ni", "I. Coast": "iv", "Senegal": "sg", "Gambia": "ga",
    "Guinea Bissau": "pu", "Sierra Leone": "sl", "Liberia": "li",
    "Guinea Conakry": "gv", "Burkina Faso": "uv", "Mali": "ml",
    "Niger": "ng", "Chad": "cd", "Cape Verde": "cv", "Togo": "to",
    "Benin": "bn", "Cameroon": "cm", "Gabon": "gb", "Congo Brazzaville": "cf",
    "Central Africa R.": "ct", "Ghana": "gh", "Namibia": "wa",
    "Mozzambique": "mz", "Madagascar": "ma", "Equatorial Guinea": "ek",
    "Sao Tome & Principe": "tp", "DRC": "cg", "Angola": "ao",
    "South Sudan": "od", "Ethiopia": "et", "Eritrea": "er",
    "Djibouti": "dj", "Somalia": "so", "Kenya": "ke", "Uganda": "ug",
    "Tanzania": "tz", "Rwanda": "rw", "Burundi": "by", "Malawi": "mi",
    "Zambia": "za", "Zimbabwe": "zi", "South Africa": "sf", "Botswana": "bc",
}

PCT_HEADERS = [
    "Population growth rate (2025 est.)", "Age structure 0-14 years",
    "Age structure 15-64 years", "Age structure >= 65 years",
    "urban population", "Literacy (Total)", "Literacy (Male)", "Literacy (Female)",
    "GDP growth rate (2024est.)", "Inflation rate (2024 est.)",
    "Mobile cellular (pourcentage population)", "Internet users (pourcentage population)",
]

INT_HEADERS = [
    "Population (2025 est.)", "Area (land): sq km", "Mobile cellular users (2019 est)",
    "Internet users (2018 est)", "Urban", "below 14years", "Below Poverty line",
]


def sync_clone() -> None:
    """Clone (ou met à jour) le dépôt public des JSON pays."""
    if (CLONE / ".git").exists():
        subprocess.run(["git", "-C", str(CLONE), "pull", "--quiet"], check=True)
        return
    CLONE.mkdir(parents=True, exist_ok=True)
    if any(CLONE.iterdir()):
        raise SystemExit(f"{CLONE} n'est pas vide et n'est pas un clone git — videz-le d'abord.")
    subprocess.run(["git", "clone", "--depth", "1", "--quiet", REPO, str(CLONE)], check=True)


def get_nested(data, *keys):
    for key in keys:
        if not isinstance(data, dict):
            return None
        data = data.get(key, {})
    return data if isinstance(data, (str, int, float)) else None


def parse_float(text):
    if text in (None, ""):
        return None
    match = re.search(r"\d+(\.\d+)?", str(text).replace(",", "").replace("%", "").strip())
    if not match:
        return None
    value = float(match.group(0))
    return value / 100.0 if "%" in str(text) else value


def parse_int(text):
    if text in (None, ""):
        return None
    match = re.search(r"\d+", str(text).replace(",", "").replace(" ", ""))
    return int(match.group(0)) if match else None


def update_frame(frame: pd.DataFrame) -> int:
    updated = 0
    for country, code in COUNTRY_CODES.items():
        path = CLONE / "africa" / f"{code}.json"
        if not path.exists():
            continue
        with path.open(encoding="utf-8") as handle:
            data = json.load(handle)
        people = data.get("People and Society", {})
        economy = data.get("Economy", {})

        values = {
            "Population (2025 est.)": parse_int(
                get_nested(people, "Population", "total", "text")
                or get_nested(people, "Population", "text")
            ),
            "Population growth rate (2025 est.)": parse_float(
                get_nested(people, "Population growth rate", "text")
            ),
            "Age structure 0-14 years": parse_float(
                get_nested(people, "Age structure", "0-14 years", "text")
            ),
            "Age structure 15-64 years": parse_float(
                get_nested(people, "Age structure", "15-64 years", "text")
            ),
            "Age structure >= 65 years": parse_float(
                get_nested(people, "Age structure", "65 years and over", "text")
            ),
            "Median age": get_nested(people, "Median age", "total", "text")
            or get_nested(people, "Median age", "text"),
            "urban population": parse_float(
                get_nested(people, "Urbanization", "urban population", "text")
            ),
            "Ethnic Groups": get_nested(people, "Ethnic groups", "text"),
            "Religions": get_nested(people, "Religions", "text"),
            "Languages": get_nested(people, "Languages", "text"),
            "Literacy (Total)": parse_float(
                get_nested(people, "Literacy", "total population", "text")
            ),
            "GDP (2024 est.) official ex rate": get_nested(
                economy, "GDP (official exchange rate)", "text"
            )
            or get_nested(
                economy,
                "Real GDP (purchasing power parity)",
                "Real GDP (purchasing power parity) 2024",
                "text",
            ),
            "GDP growth rate (2024est.)": parse_float(
                get_nested(economy, "Real GDP growth rate", "Real GDP growth rate 2024", "text")
            ),
            "GDP per capita (2019 est.)": get_nested(
                economy, "Real GDP per capita", "Real GDP per capita 2024", "text"
            ),
            "Inflation rate (2024 est.)": parse_float(
                get_nested(
                    economy,
                    "Inflation rate (consumer prices)",
                    "Inflation rate (consumer prices) 2024",
                    "text",
                )
            ),
            "Unemployment rate": get_nested(
                economy, "Unemployment rate", "Unemployment rate 2024", "text"
            ),
        }

        index = frame.index[frame["Country"] == country]
        if len(index) == 0:
            continue
        row = index[0]
        for column, value in values.items():
            if value not in (None, ""):
                frame.loc[row, column] = value

        # Volumes absolus recalculés (utilisés par les agrégats du dashboard).
        population = values["Population (2025 est.)"]
        if population:
            if values["urban population"] is not None:
                frame.loc[row, "Urban"] = round(population * values["urban population"])
            if values["Age structure 0-14 years"] is not None:
                frame.loc[row, "below 14years"] = round(
                    population * values["Age structure 0-14 years"]
                )
        updated += 1
    return updated


def add_clusters(frame: pd.DataFrame) -> pd.DataFrame:
    def main_cluster(country):
        if country == "Nigeria":
            return "NG"
        return None if country in ("India Ocean", "Islands") else "RoSSA"

    def sub_cluster(country):
        if country == "Nigeria":
            return "NG"
        if country in ("India Ocean", "Islands"):
            return None
        return "WA" if country in WA_COUNTRIES else "EESA"

    frame["Main_Cluster"] = frame["Country"].apply(main_cluster)
    frame["Sub_Cluster"] = frame["Country"].apply(sub_cluster)
    # Les deux clés business viennent juste après le pays.
    columns = ["Country", "Main_Cluster", "Sub_Cluster"]
    return frame[columns + [c for c in frame.columns if c not in columns]]


def write_workbook(frame: pd.DataFrame) -> None:
    book = openpyxl.Workbook()
    sheet = book.active
    sheet.title = "Factbook FMCG 2026"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    edge = Side(border_style="thin", color="D3D3D3")
    border = Border(left=edge, right=edge, top=edge, bottom=edge)

    for position, title in enumerate(frame.columns, 1):
        cell = sheet.cell(row=1, column=position, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
        sheet.column_dimensions[get_column_letter(position)].width = 25

    for row_index, row in enumerate(dataframe_to_rows(frame, index=False, header=False), 2):
        for column_index, value in enumerate(row, 1):
            cell = sheet.cell(row=row_index, column=column_index, value=value)
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            column = frame.columns[column_index - 1]
            if isinstance(value, (int, float)):
                if column in PCT_HEADERS:
                    cell.number_format = "0.00%"
                elif column in INT_HEADERS:
                    cell.number_format = "#,##0"

    sheet.auto_filter.ref = sheet.dimensions
    sheet.freeze_panes = "B2"
    book.save(TARGET_XLSX)


def main() -> None:
    sync_clone()
    frame = pd.read_excel(SOURCE_XLSX)
    updated = update_frame(frame)
    frame = add_clusters(frame)
    write_workbook(frame)
    frame.to_json(TARGET_JSON, orient="records", indent=4, force_ascii=False)
    print(f"{updated} pays synchronisés -> {TARGET_XLSX.name} + {TARGET_JSON.name}")
    build_json()


if __name__ == "__main__":
    main()
